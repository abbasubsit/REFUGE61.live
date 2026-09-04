# -*- coding: utf-8 -*-
"""
Compare the HOME sheet of Bjørn's .xlsm against what is currently published.

He says he "made a few changes in the first page Home" without saying which
language, so all three columns are checked: English against the generated
copy modules' `en` field (which mirrors the JSX), French against
copy.fr.ts, Danish against copy.da.ts.
"""
import io
import os
import re

import openpyxl

SRC = os.path.join(
    r"C:\Users\Abbas\OneDrive\Desktop\noman",
    "REFUGE61_Website_V3_FR_corrige_CHANGES_MAJ.xlsm",
)
OUT = "_scripts/i18n/_home_diff.txt"


def norm(s):
    return re.sub(r"\s+", " ", (s or "")).strip()


def published(path, field):
    """{T-id: value} from a generated copy module."""
    src = io.open(path, encoding="utf-8").read()
    out = {}
    for m in re.finditer(
        r'(T\d{3}): \{.*?en: "((?:[^"\\]|\\.)*)",\s*tr: "((?:[^"\\]|\\.)*)"',
        src, re.S,
    ):
        value = m.group(2) if field == "en" else m.group(3)
        out[m.group(1)] = value.replace('\\"', '"').replace("\\\\", "\\")
    return out


def main():
    wb = openpyxl.load_workbook(SRC, keep_vba=True, data_only=True)
    if "HOME" not in wb.sheetnames:
        raise SystemExit("no HOME sheet; sheets are: %s" % wb.sheetnames)

    live = {
        "en": published("lib/content/copy.fr.ts", "en"),
        "fr": published("lib/content/copy.fr.ts", "tr"),
        "da": published("lib/content/copy.da.ts", "tr"),
    }

    lines, counts = [], {"en": 0, "fr": 0, "da": 0}
    rows = 0
    for r in wb["HOME"].iter_rows(min_row=4, values_only=True):
        if not r or not r[0]:
            continue
        tid = str(r[0]).strip()
        if not re.fullmatch(r"T\d{3}", tid):
            continue
        rows += 1
        incoming = {"en": norm(r[2]), "fr": norm(r[3])}
        incoming["da"] = norm(r[4]) if len(r) > 4 else ""

        for lang in ("en", "fr", "da"):
            new = incoming[lang]
            old = norm(live[lang].get(tid, ""))
            if not new or not old or new == old:
                continue
            counts[lang] += 1
            lines.append("%s  [%s]\n   live: %s\n   new : %s\n" % (tid, lang.upper(), old, new))

    header = ("HOME rows: %d | changed -- EN: %d  FR: %d  DA: %d"
              % (rows, counts["en"], counts["fr"], counts["da"]))
    io.open(OUT, "w", encoding="utf-8").write(header + "\n\n" + "\n".join(lines))
    print(header)


if __name__ == "__main__":
    main()
