# -*- coding: utf-8 -*-
"""
Dump the client's second French revision pass.

He marks revisions by font colour ("Last changes in BLUE"), so the changed
cells are found by inspecting each cell's colour rather than by diffing
against the previous file -- a diff would also surface his first round of
corrections, which are already live.
"""
import io
import os

import openpyxl

SRC = os.path.join(
    r"C:\Users\Abbas\OneDrive\Desktop\noman",
    "REFUGE61_Website_V2_EN_FR_DA_FR_corrige.xlsx",
)

wb = openpyxl.load_workbook(SRC)
out = ["SHEETS: %s" % ", ".join(wb.sheetnames)]

for ws in wb.worksheets:
    out.append("")
    out.append("===== %s (%d rows) =====" % (ws.title, ws.max_row))
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            if cell.value is None:
                continue
            colour = cell.font.color
            rgb = getattr(colour, "rgb", None) if colour else None
            if not isinstance(rgb, str):
                continue
            # Skip black / automatic / the red of the earlier round.
            if rgb.upper() in ("FF000000", "00000000"):
                continue
            ident = ws.cell(row=cell.row, column=1).value
            out.append(
                "%s | %s | col%d | %s | %s"
                % (ws.title, ident, cell.column, rgb, str(cell.value)[:200])
            )

io.open("_scripts/i18n/_v2_changes.txt", "w", encoding="utf-8").write("\n".join(out))
print("wrote %d lines" % len(out))
