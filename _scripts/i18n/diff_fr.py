# -*- coding: utf-8 -*-
"""
Diff the client's newest French spreadsheet against the one already applied.

He said "last changes in BLUE", but the file carries no blue font run -- its
only coloured cells are the reds of the round already live, and its CHANGES FR
sheet documents an older pass. So the new edits are found by comparing the
French column of the two workbooks, ID by ID, which does not depend on how the
changes were marked.
"""
import io
import os

import openpyxl

DESKTOP = r"C:\Users\Abbas\OneDrive\Desktop\noman"
# What is actually live: the snapshot committed alongside the French build.
APPLIED_JSON = "_scripts/i18n/_client_fr.json"
# Point this at whatever the client sends next.
NEW = os.path.join(DESKTOP, "REFUGE61_Website_V2_EN_FR_DA_FR_corrige.xlsx")

SHEETS = ["HOME", "THE LODGE", "PRACTICAL INFO", "LET'S TALK", "TERMS", "NAVIGATION"]


def french(path):
    wb = openpyxl.load_workbook(path)
    out = {}
    for name in SHEETS:
        if name not in wb.sheetnames:
            continue
        for row in wb[name].iter_rows(min_row=4, values_only=True):
            if not row or not row[0]:
                continue
            out[str(row[0]).strip()] = {
                "sheet": name,
                "en": (row[2] or "").strip(),
                "fr": (row[3] or "").strip(),
            }
    return out


def applied():
    import json
    data = json.load(io.open(APPLIED_JSON, encoding="utf-8"))
    out = {}
    for sheet, rows in data["sheets"].items():
        for r in rows:
            out[str(r["id"]).strip()] = {
                "sheet": sheet,
                "en": (r["en"] or "").strip(),
                "fr": (r["fr"] or "").strip(),
            }
    return out


old, new = applied(), french(NEW)

lines = []
changed = added = removed = 0
for key in sorted(set(old) | set(new), key=lambda k: (len(k), k)):
    a, b = old.get(key), new.get(key)
    if a and b and a["fr"] != b["fr"]:
        changed += 1
        lines.append("CHANGED %s (%s)" % (key, b["sheet"]))
        lines.append("   EN  : %s" % b["en"][:150])
        lines.append("   was : %s" % a["fr"])
        lines.append("   now : %s" % b["fr"])
        lines.append("")
    elif b and not a and b["fr"]:
        added += 1
        lines.append("ADDED   %s (%s): %s" % (key, b["sheet"], b["fr"][:150]))
    elif a and not b and a["fr"]:
        removed += 1
        lines.append("REMOVED %s (%s): %s" % (key, a["sheet"], a["fr"][:150]))

header = "changed: %d   added: %d   removed: %d   (old %d ids, new %d ids)" % (
    changed, added, removed, len(old), len(new))
io.open("_scripts/i18n/_fr_diff.txt", "w", encoding="utf-8").write(
    header + "\n\n" + "\n".join(lines))
print(header)
