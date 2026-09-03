# -*- coding: utf-8 -*-
"""Survey the client's newest workbook: Danish coverage and any comments."""
import io
import os

import openpyxl

SRC = os.path.join(
    r"C:\Users\Abbas\OneDrive\Desktop\noman",
    "REFUGE61_Website_V3_FR_corrige_CHANGES_MAJ (1).xlsx",
)

wb = openpyxl.load_workbook(SRC)
out = ["FILE: %s" % os.path.basename(SRC), "SHEETS: %s" % ", ".join(wb.sheetnames), ""]

total = {"rows": 0, "en": 0, "fr": 0, "da": 0}
comments = []
extra_cols = set()

for ws in wb.worksheets:
    n = {"rows": 0, "en": 0, "fr": 0, "da": 0}
    for row in ws.iter_rows(min_row=4):
        ident = row[0].value if row else None
        if not ident:
            continue
        n["rows"] += 1
        vals = [c.value for c in row]
        if len(vals) > 2 and vals[2]:
            n["en"] += 1
        if len(vals) > 3 and vals[3]:
            n["fr"] += 1
        if len(vals) > 4 and vals[4]:
            n["da"] += 1
        # Anything past the Danish column is new -- a comments column, maybe.
        for c in row[5:]:
            if c.value:
                extra_cols.add(c.column)
                comments.append("%s | %s | col%d | %s"
                                % (ws.title, ident, c.column, str(c.value)[:300]))
    for c in ws._comments if hasattr(ws, "_comments") else []:
        comments.append("CELL COMMENT %s | %s" % (ws.title, c.text[:300]))
    out.append("%-16s rows %3d | EN %3d | FR %3d | DA %3d"
               % (ws.title, n["rows"], n["en"], n["fr"], n["da"]))
    for k in total:
        total[k] += n[k]

out.append("")
out.append("TOTAL            rows %3d | EN %3d | FR %3d | DA %3d"
           % (total["rows"], total["en"], total["fr"], total["da"]))
out.append("")
out.append("columns beyond Danish: %s" % (sorted(extra_cols) or "none"))
out.append("comment/extra cells: %d" % len(comments))
out.extend(comments[:80])

io.open("_scripts/i18n/_v3_survey.txt", "w", encoding="utf-8").write("\n".join(out))
print("wrote %d lines" % len(out))
